from openpyxl import load_workbook
file = "C:/Fakhoury_DataMining/Fakhoury_Field_Descriptions.xlsx"



def get_cell_values():
	wb = load_workbook(filename = file,use_iterators=True)
	ws = wb['Sheet1']
	result = []
	range = "B23:B106"
	for row in ws.iter_rows(range_string=range):
		for cell in row:
			value = cell.value	
			result.append(value)
	return result

	




def make_dict(somelist):
	climate_dict = {}
	for i in somelist:
		t = str(i)
		if "30" in t:
			climate_dict[t] = "{0} from 30 year average".format(t[:t.find("_")])
		elif "Tmax_2" in t:
			climate_dict[t] = "Maximum temperature in {0}".format(t[-4:])
		elif "Tmax_" in t:
			climate_dict[t] = "Maximum temperature in {0} {1}".format(t[5:-5],t[-4:])
		elif "Tmin_2" in t:
			climate_dict[t] = "Minimum temperature in {0}".format(t[-4:])
		elif "Tmin_" in t:
			climate_dict[t] = "Minimum temperature in {0} {1}".format(t[5:-5],t[-4:])
		elif "Tmean_2" in t:
			climate_dict[t] = "Average temperature in {0}".format(t[-4:])
		elif "Tmean_" in t:
			climate_dict[t] = "Average temperature in {0} {1}".format(t[6:-5],t[-4:])
		elif "Precip_2" in t:
			climate_dict[t] = "Total precipitation in {0}".format(t[-4:])
		elif "Precip_" in t:
			climate_dict[t] = "Total precipitation for {0} {1}".format(t[7:-5],t[-4:])
	return climate_dict

def write_dict_to_excel(somedict):
	wb = load_workbook(filename=file)
	ws = wb['Sheet1']
	for i in range(23,107):
		cell = ws.cell(column=2,row = i)
		desc = ws.cell(column=3,row = i)
		units = ws.cell(column=4,row=i)
		desc.value = somedict[str(cell.value)]
		if "Precip" in str(cell.value):
			units.value = "millimeters"
		else:
			units.value = "degrees C"
	wb.save(file)
	

# print get_cell_values()
# print make_dict(get_cell_values())
write_dict_to_excel(make_dict(get_cell_values()))
